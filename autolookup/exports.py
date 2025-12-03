from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table as PDFTable,
    TableStyle,
)
from openpyxl.styles import PatternFill
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime
import rich
from rich import print
import pandas as pd

from historyUtils import load_history
from log import logger



def export_document(vin: str, data: dict):
    try:
        date = __import__('datetime').datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        filename = f"{vin}_data.txt"
        with open(filename, 'w') as f:
            f.write(f"Date requested: {date}\n")
            for key, value in data.items():
                f.write(f"{key}: {value}\n")
                
        print(f"[green]VIN data exported to {filename}[/green]")
    except Exception as e:
        logger.error(f"Error exporting VIN data to TXT: {e}")
        print("[red]Error exporting VIN data to TXT.[/red]")
        return
def export_pdf(vin: str, data: dict):
    filename = f"{vin}_data.pdf"
    try: 
        # Create PDF document
        doc = SimpleDocTemplate(
            filename,
            pagesize=letter,
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=18,
        )

        styles = getSampleStyleSheet()
        story = []

        # Title
        title = f"<para alignment='center'><b>VIN Report: {vin}</b></para>"
        story.append(Paragraph(title, styles["Title"]))
        story.append(Spacer(1, 12))

        # Timestamp
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        story.append(Paragraph(f"<b>Date Generated:</b> {timestamp}", styles["Normal"]))
        story.append(Spacer(1, 12))

        # VIN Data Table
        table_data = [["Data", "Value"]]
        for key, value in data.items():
            table_data.append([key, str(value)])

        table = PDFTable(table_data, colWidths=[150, 350])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 12),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
            ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))

        story.append(table)

        # Build PDF
        doc.build(story)

        print(f"[green]PDF exported to {filename}[/green]")
        return 
    except Exception as e:
        logger.error(f"Error exporting VIN data to PDF: {e}")
        print("[red]Error exporting VIN data to PDF.[/red]")
        return   
    

def export_batch_pdf(all_results: list):
    try:
        filename = f"batch_vin_lookup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table as PDFTable, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        doc = SimpleDocTemplate(filename, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
        styles = getSampleStyleSheet()
        story = []

        story.append(Paragraph(f"<b>Batch VIN Lookup Report</b>", styles["Title"]))
        story.append(Spacer(1, 12))
        for entry in all_results:
            vin = entry['vin']
            data = entry['data']
            table_data = [["Field", "Value"]]
            for key, value in data.items():
                table_data.append([key, str(value)])
            table = PDFTable(table_data, colWidths=[150, 350])
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 12),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
                ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            story.append(Paragraph(f"<b>VIN: {vin}</b>", styles["Heading3"]))
            story.append(table)
            story.append(Spacer(1, 12))

        doc.build(story)
        print(f"[green]All results exported to {filename}[/green]")
    except Exception as e:
        logger.error(f"Error exporting batch PDF: {e}")
        print("[red]Error exporting batch PDF.[/red]")
def export_batch_txt(all_results: list):
    try:
        filename = f"batch_vin_lookup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        with open(filename, 'w') as f:
            for entry in all_results:
                vin = entry['vin']
                data = entry['data']
                f.write(f"VIN: {vin}\n")
                for key, value in data.items():
                    f.write(f"{key}: {value}\n")
                f.write("\n")
        print(f"[green]All results exported to {filename}[/green]")
    except Exception as e:
        logger.error(f"Error exporting batch TXT: {e}")
        print("[red]Error exporting batch TXT.[/red]")
def export_batch_excel(all_results: list):
    try:

        filename = f"batch_vin_lookup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        all_data = []
        for entry in all_results:
            vin = entry['vin']
            data = entry['data']
            data_row = {"VIN": vin}
            data_row.update(data)
            all_data.append(data_row)
        df = pd.DataFrame(all_data)
        df.to_excel(filename, index=False)
        print(f"[green]All results exported to {filename}[/green]")
    except Exception as e:
        logger.error(f"Error exporting batch Excel: {e}")
        print("[red]Error exporting batch Excel.[/red]")

def export_history_to_excel():
    history = load_history()
    if not history:
        print("[yellow]No history to export.[/yellow]")
        return

    try:
        filename = f"vin_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        all_data = []
        for entry in history:
            vin = entry.get("vin") or "N/A"
            data = entry.get("data") or {}
            data_row = {"VIN": vin}
            data_row.update(data)
            all_data.append(data_row)

        df = pd.DataFrame(all_data)
        df.to_excel(filename, index=False)
        print(f"[green]History exported to {filename}[/green]")
    except Exception as e:
        logger.error(f"Error exporting history to Excel: {e}")
        print("[red]Error exporting history to Excel.[/red]")
def export_history_to_txt():
    history = load_history()
    if not history:
        print("[yellow]No history to export.[/yellow]")
        return

    try:
        filename = f"vin_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        with open(filename, 'w') as f:
            for entry in history:
                vin = entry.get("vin") or "N/A"
                data = entry.get("data") or {}
                f.write(f"VIN: {vin}\n")
                for key, value in data.items():
                    f.write(f"{key}: {value}\n")
                f.write("\n")
        print(f"[green]History exported to {filename}[/green]")
    except Exception as e:
        logger.error(f"Error exporting history to TXT: {e}")
        print("[red]Error exporting history to TXT.[/red]")
def export_history_to_pdf():
    history = load_history()
    if not history:
        print("[yellow]No history to export.[/yellow]")
        return

    try:
        filename = f"vin_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table as PDFTable, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        doc = SimpleDocTemplate(filename, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
        styles = getSampleStyleSheet()
        story = []

        story.append(Paragraph(f"<b>VIN History Report</b>", styles["Title"]))
        story.append(Spacer(1, 12))
        for entry in history:
            vin = entry.get("vin") or "N/A"
            data = entry.get("data") or {}
            table_data = [["Field", "Value"]]
            for key, value in data.items():
                table_data.append([key, str(value)])
            table = PDFTable(table_data, colWidths=[150, 350])
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 12),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
                ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            story.append(Paragraph(f"<b>VIN: {vin}</b>", styles["Heading3"]))
            story.append(table)
            story.append(Spacer(1, 12))

        doc.build(story)
        print(f"[green]History exported to {filename}[/green]")
    except Exception as e:
        logger.error(f"Error exporting history to PDF: {e}")
        print("[red]Error exporting history to PDF.[/red]")

def export_comparison_excel(vin1_data: dict, vin2_data: dict, vin1: str, vin2: str):
    try:
        filename = f"vin_comparison_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Comparison"

        # Header row
        headers = ["Field", vin1, vin2]
        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="D9D9D9")
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # Style header
        for col in range(1, 4):
            cell = ws.cell(row=1, column=col)
            cell.font = bold_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border

        # Populate table
        all_keys = sorted(set(vin1_data.keys()).union(vin2_data.keys()))

        for row_idx, key in enumerate(all_keys, start=2):
            val1 = vin1_data.get(key, "")
            val2 = vin2_data.get(key, "")

            ws.append([key, val1, val2])

            # Style each row
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left")

            # Highlight differences
            if val1 != val2:
                diff_fill = PatternFill("solid", fgColor="FFF2CC")  # light yellow

                ws.cell(row=row_idx, column=1).fill = diff_fill
                ws.cell(row=row_idx, column=2).fill = diff_fill
                ws.cell(row=row_idx, column=3).fill = diff_fill

                ws.cell(row=row_idx, column=2).font = Font(color="FF0000")
                ws.cell(row=row_idx, column=3).font = Font(color="FF0000")

            # Add borders
            for col in range(1, 4):
                ws.cell(row=row_idx, column=col).border = border

        # Auto column widths
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        # Freeze header row
        ws.freeze_panes = "A2"

        wb.save(filename)
        print(f"[green]Excel comparison exported to {filename}[/green]")

    except Exception as e:
        print(f"[red]Error exporting Excel comparison: {e}[/red]")
    except Exception as e:
        print(f"[red]Error exporting comparison Excel: {e}[/red]")
def export_comparison_txt(vin1_data: dict, vin2_data: dict, vin1: str, vin2: str):
    filename = f"VIN_comparison_{vin1}_{vin2}.txt"
    with open(filename, "w") as f:
        f.write(f"VIN Comparison: {vin1} vs {vin2}\n\n")
        all_keys = sorted(set(list(vin1_data.keys()) + list(vin2_data.keys())))
        for key in all_keys:
            val1 = vin1_data.get(key, "N/A")
            val2 = vin2_data.get(key, "N/A")
            f.write(f"{key}: {val1} | {val2}\n")
    print(f"[green]Comparison exported to {filename}[/green]")
def export_comparison_pdf(vin1_data: dict, vin2_data: dict, vin1: str, vin2: str):
    try:
        filename = f"vin_comparison_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        doc = SimpleDocTemplate(filename, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
        styles = getSampleStyleSheet()
        story = []

        story.append(Paragraph(f"<b>VIN Comparison Report</b>", styles["Title"]))
        story.append(Spacer(1, 12))

        all_keys = sorted(set(vin1_data.keys()).union(vin2_data.keys()))
        table_data = [["Field", vin1, vin2]]

        for key in all_keys:
            val1 = vin1_data.get(key, "")
            val2 = vin2_data.get(key, "")
            table_data.append([key, val1, val2])

        table = PDFTable(table_data, colWidths=[150, 200, 200])
        # Style with highlighting differences
        style = TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
            ("ALIGN", (0,0), (-1,-1), "LEFT"),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ])

        # Highlight differences
        for i, row in enumerate(table_data[1:], start=1):
            if row[1] != row[2]:
                style.add("TEXTCOLOR", (1,i), (2,i), colors.red)

        table.setStyle(style)
        story.append(table)
        doc.build(story)
        print(f"[green]Comparison PDF exported to {filename}[/green]")
        logger.info(f"Comparison PDF exported to {filename}")

    except Exception as e:
        print(f"[red]Error exporting comparison PDF: {e}[/red]")
        logger.error(f"Error exporting comparison PDF: {e}")